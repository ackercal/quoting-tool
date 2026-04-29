"""Generate Excel matching the original Claude Version.xlsx Forecast sheet style."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from calculations import (
    LABOR_HOURS_FORMED_PARTS, LABOR_HOURS_CUSTOM_AUTO,
    PART_HOURS_FORMED_PARTS, PART_HOURS_CUSTOM_AUTO,
    PROJECT_HOURS,
    ROBOT_IMPROVEMENT,
    TRIAL_REDUCTION,
)

YEARS = [2026, 2027, 2028]
ROLES = ["RPE", "ME", "Tech"]

# Light blue fill matching Excel theme:5 data cells in original
BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")

F11 = Font(name="Calibri", size=11)
F11B = Font(name="Calibri", size=11, bold=True)


def v(ws, row, col, value, bold=False, fill=None, align=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = F11B if bold else F11
    if fill:
        c.fill = fill
    if align:
        c.alignment = Alignment(horizontal=align)
    return c


def write_sheet(ws, set_name, labor_set, part_set):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9

    row = 1

    # Year header row
    v(ws, row, 2, "2026")
    v(ws, row, 3, "2027")
    v(ws, row, 4, "2028")
    row += 1

    # ── Labor Hours Per Operation ─────────────────────────────────────────────
    v(ws, row, 1, "Labor Hours Per Operation", bold=True)
    row += 1

    op_items = [
        ("Forming — Pre-IF",        "pre_if_forming"),
        ("Forming — IF",            "if_forming"),
        ("Forming — Duplicate",     "dup_forming"),
        ("Scanning — First",        "first_scan"),
        ("Scanning — Duplicate",    "dup_scan"),
        ("Cutting — First",         "first_cut"),
        ("Cutting — Duplicate",     "dup_cut"),
    ]

    for op_label, op_key in op_items:
        if op_key not in labor_set:
            continue
        v(ws, row, 1, op_label, bold=True)
        row += 1
        for role in ROLES:
            v(ws, row, 1, role, align="left")
            for ci, yr in enumerate(YEARS, 2):
                val = labor_set[op_key].get(yr, labor_set[op_key].get(2026, {})).get(role, 0)
                v(ws, row, ci, val, fill=BLUE_FILL)
            row += 1

    row += 1  # blank row between sections

    # ── Robot Hour Improvement ────────────────────────────────────────────────
    v(ws, row, 1, "Robot Hour Improvement (from 2026)", bold=True)
    v(ws, row, 2, "2026")
    v(ws, row, 3, "2027")
    v(ws, row, 4, "2028")
    row += 1

    for cat_label, cat_key in [("Forming", "forming"), ("Scanning", "scanning"), ("Cutting", "cutting")]:
        v(ws, row, 1, cat_label)
        for ci, yr in enumerate(YEARS, 2):
            v(ws, row, ci, ROBOT_IMPROVEMENT[cat_key].get(yr, 1.0), fill=BLUE_FILL)
        row += 1

    row += 1

    # ── Labor Hours Per Part ──────────────────────────────────────────────────
    v(ws, row, 1, "Labor Hours Per Part", bold=True)
    v(ws, row, 2, "2026")
    v(ws, row, 3, "2027")
    v(ws, row, 4, "2028")
    row += 1

    part_items = [
        ("Prep for Shipping (Tech)",   "palletize_tech",    True),
        ("Unistrut (Tech)",            "unistrut_tech",     True),
        ("Purchaser Setup",            "purchaser_setup",   False),
        ("PM Setup",                   "pm_setup",          False),
        ("Purchaser Overhead",         "purchaser_overhead",False),
        ("PM Overhead",                "pm_overhead",       False),
    ]

    for label, key, year_varying in part_items:
        v(ws, row, 1, label)
        val = part_set.get(key, 0)
        for ci, yr in enumerate(YEARS, 2):
            cell_val = val.get(yr, 0) if isinstance(val, dict) else val
            v(ws, row, ci, cell_val, fill=BLUE_FILL)
        row += 1

    row += 1

    # ── Labor Hours Per Project ───────────────────────────────────────────────
    v(ws, row, 1, "Labor Hours Per Project", bold=True)
    v(ws, row, 2, "2026")
    v(ws, row, 3, "2027")
    v(ws, row, 4, "2028")
    row += 1

    v(ws, row, 1, "Overhead")
    row += 1

    project_items = [
        ("Purchaser", "purchaser"),
        ("Project Manager", "pm"),
    ]
    for label, key in project_items:
        v(ws, row, 1, label)
        for ci, yr in enumerate(YEARS, 2):
            v(ws, row, ci, PROJECT_HOURS[key].get(yr, 0), fill=BLUE_FILL)
        row += 1

    row += 1

    # ── Trial Reduction ───────────────────────────────────────────────────────
    v(ws, row, 1, "Trial Reduction", bold=True)
    v(ws, row, 2, "2026")
    v(ws, row, 3, "2027")
    v(ws, row, 4, "2028")
    row += 1

    v(ws, row, 1, "Pre-IF & IF procedures × factor, rounded up")
    for ci, yr in enumerate(YEARS, 2):
        v(ws, row, ci, TRIAL_REDUCTION.get(yr, 1.0), fill=BLUE_FILL)
    row += 1


# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

for name, labor, part in [
    ("Formed Parts", LABOR_HOURS_FORMED_PARTS, PART_HOURS_FORMED_PARTS),
    ("Custom Auto",  LABOR_HOURS_CUSTOM_AUTO,  PART_HOURS_CUSTOM_AUTO),
]:
    ws = wb.create_sheet(title=name)
    write_sheet(ws, name, labor, part)

out = r"C:\Users\CalvinAcker\Downloads\labor_forecasts.xlsx"
wb.save(out)
print(f"Saved: {out}")
