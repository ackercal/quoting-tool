"""Inspect Forecast sheet and check which cells reference C17 (year)."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl, re

EXCEL_PATH = r"C:\Users\CalvinAcker\Downloads\claude_version2.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

# ── Forecast sheet ─────────────────────────────────────────────────────────────
print("=== Forecast sheet ===")
ws_f = wb["Forecast"]
for row in ws_f.iter_rows(min_row=1, max_row=60):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate:6s} | {repr(cell.value)[:80]}")

# ── Which cells in Labor+Robot reference C17 ──────────────────────────────────
print("\n=== Cells that reference C17 (Year) ===")
ws = wb["Labor + Robot Time"]
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and "C17" in cell.value:
            print(f"  {cell.coordinate}: {cell.value}")

# ── Which cells reference C41 (unistrut toggle) ───────────────────────────────
print("\n=== Cells that reference C41 (Unistrut toggle) ===")
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and "C41" in cell.value:
            print(f"  {cell.coordinate}: {cell.value}")

# ── Robot improvement: check if any formula applies improvement factor ─────────
print("\n=== P26 and P27 formulas ===")
print(f"  P26: {ws['P26'].value}")
print(f"  P27: {ws['P27'].value}")

print("\n=== Robot-related rows (F9,F15,F21,F27 = robot time inputs) ===")
for c in ["F9","G9","F15","G15","F21","G21","F27","G27"]:
    print(f"  {c}: {ws[c].value}")
